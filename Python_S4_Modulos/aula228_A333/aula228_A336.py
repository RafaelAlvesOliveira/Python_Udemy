# openpyxl - ler e alterar dados de uma planilha
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import load_workbook    # type: ignore[import]
# from openpyxl.cell import Cell    # type: ignore[import]
# from openpyxl.worksheet.worksheet import Worksheet    # type: ignore[import]
import os
import sys

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Garante que estamos usando um nome único para o arquivo de saída
NEW_WORKBOOK_PATH = ROOT_FOLDER / 'workbook_atualizado_temp.xlsx'

# Verifica se o arquivo de saída já existe e tenta removê-lo
if NEW_WORKBOOK_PATH.exists():
    try:
        os.unlink(NEW_WORKBOOK_PATH)
    except PermissionError:
        print(f"Erro: Não foi possível remover \n"
              f"o arquivo existente {NEW_WORKBOOK_PATH}")
        sys.exit(1)


try:
    workbook = load_workbook(WORKBOOK_PATH)
    worksheet = workbook['Minha planilha']

    # Suas modificações aqui
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if cell.value == 'Maria':
                cell.offset(column=1).value = 23

    # worksheet['B3'].value = 14

    # Tenta salvar o arquivo
    workbook.save(NEW_WORKBOOK_PATH)
    print(f"Arquivo salvo com sucesso em: {NEW_WORKBOOK_PATH}")

    # Renomeia para o nome final (opcional)
    final_path = ROOT_FOLDER / 'workbook_atualizado.xlsx'
    try:
        if final_path.exists():
            os.unlink(final_path)
        os.rename(NEW_WORKBOOK_PATH, final_path)
        print(f"Arquivo final disponível em: {final_path}")
    except Exception as e:
        print(f"AVISO: Não foi possível renomear o arquivo: {str(e)}")
        print(f"O arquivo foi mantido como: {NEW_WORKBOOK_PATH}")

except Exception as e:
    print(f"Erro durante o processamento: {str(e)}")
finally:
    if 'workbook' in locals():
        workbook.close()
